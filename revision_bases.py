#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# revision_bases.py - Modulo de Revision de Bases ITS/TAR/PrEP
# Proposito: Automatizar el proceso manual de 6 pasos descrito en
#   "Proceso de Revision de EESS, BMU, OBC"
# Fuente: Trama Unificada (71 columnas, 3 origenes: ITS, TAR, PrEP)
# Estilo: NVIH develop (lista plana de diccionarios, sin decoradores)
# Creado: 2026-05-25
# Dependencias: pandas, openpyxl

import os
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import re
import logging
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
log = logging.getLogger(__name__)

# ──────────────────────────────────────────────
# CONFIGURACION: columnas esperadas
# ──────────────────────────────────────────────
COLUMNAS_ITS = [
    'sector_its', 'diresa_its', 'region_its', 'provincia_its', 'distrito_its',
    'eess_its', 'renipres_its', 'tipo_oferta_its', 'doc_brigadista_its',
    'nom_brigadista_its', 'ape_brigadista_its', 'brigada_its',
    'fecha_intervencion', 'tipo_lugar_its',
    'uid', 'edad', 'sexo', 'tipo_poblacion', 'nacionalidad',
    'fecha_consejeria', 'fecha_tamizaje', 'tipo_tamizaje', 'resultado',
    'vinculo_tipo', 'vinculo_estado', 'vinculo_eess_dest',
    'vinculo_fecha', 'vinculo_fecha_reg', 'fecha_reg_its', 'fecha_mod_its',
    'origen_registro'
]

COLUMNAS_TAR = [
    'sector_vih', 'diris_vih', 'depto_vih', 'provincia_vih', 'distrito_vih',
    'eess_vih', 'renipres_vih', 'doc_vih', 'nombre_vih', 'apellido_vih',
    'condicion_vih', 'fecha_inicio_tar', 'fecha_ult_atencion',
    'fecha_prox_cita', 'fecha_ult_abandono', 'fecha_recup_abandono',
    'fecha_derivacion', 'lugar_derivacion', 'fecha_creacion_vih', 'fecha_mod_vih'
]

COLUMNAS_PREP = [
    'sector_prep', 'diris_prep', 'depto_prep', 'provincia_prep', 'distrito_prep',
    'eess_prep', 'renipres_prep', 'doc_prep', 'nombre_prep', 'apellido_prep',
    'fecha_inicio_prep', 'eess_inicio_prep', 'modalidad_prep',
    'fecha_reingreso_prep', 'condicion_actual', 'fecha_ult_atencion_prep',
    'fecha_ult_tamizaje_vih', 'resultado_ult_tamizaje_vih',
    'fecha_creacion_prep', 'fecha_mod_prep'
]

COLUMNAS_ESPERADAS = COLUMNAS_ITS + COLUMNAS_TAR + COLUMNAS_PREP

# ──────────────────────────────────────────────
# VERIFICACIONES
# ──────────────────────────────────────────────
VERIFICACIONES = [
    {
        'n': 'filtro_vih_dvi',
        't': 'Filtro: solo registros VIH y DVI',
        'd': 'Registros con tipo de tamizaje VIH o DVI (Dual VIH). '
             'Se excluyen HEB (Hepatitis B), SIF (Sifilis), DSI (Dual Sifilis) '
             'y HEC (Hepatitis C). La columna tipo_tamizaje se resalta en amarillo '
             'en el Excel descargado.',
        'c': ['tipo_tamizaje'],
        'resaltar': 'tipo_tamizaje',
        'p': 'baja',
        'cat': 'filtro'
    },
    {
        'n': 'duplicados_tamizaje',
        't': 'Duplicados por UID',
        'd': 'Identifica registros con mismo UID duplicado',
        'c': ['uid'],
        'resaltar': 'uid',
        'p': 'media',
        'cat': 'duplicados'
    },
    {
        'n': 'reactivos_sin_vinculacion',
        't': 'Reactivos VIH sin vinculacion efectiva',
        'd': 'Pacientes con resultado REACTIVO en tamizaje VIH/Dual '
             'pero sin un registro de vinculacion efectiva a TAR o PrEP.',
        'c': ['resultado', 'vinculo_estado'],
        'resaltar': 'resultado',
        'p': 'critica',
        'cat': 'brecha'
    },
    {
        'n': 'vinculados_sin_padron',
        't': 'Vinculados que no aparecen en TAR ni PrEP',
        'd': 'Paciente con vinculacion efectiva registrada en ITS, '
             'pero sin datos correspondientes en TAR (condicion_vih) '
             'ni en PrEP (condicion_actual).',
        'c': ['vinculo_estado', 'condicion_vih', 'condicion_actual'],
        'resaltar': 'vinculo_estado',
        'p': 'alta',
        'cat': 'brecha'
    },
    {
        'n': 'reactivos_vih_sin_tar',
        't': 'Reactivos VIH sin inicio de TAR',
        'd': 'Paciente con resultado REACTIVO en tamizaje VIH '
             'pero sin fecha de inicio de TAR registrada.',
        'c': ['resultado', 'fecha_inicio_tar'],
        'resaltar': 'resultado',
        'p': 'critica',
        'cat': 'brecha'
    },
    {
        'n': 'en_tar_sin_tamizaje_its',
        't': 'En TAR sin tamizaje ITS registrado',
        'd': 'Paciente que aparece en TAR (con condicion_vih y fecha_inicio_tar) '
             'pero no tiene ningun registro en ITS (uid no aparece en datos ITS).',
        'c': ['uid', 'fecha_tamizaje', 'condicion_vih'],
        'resaltar': 'condicion_vih',
        'p': 'alta',
        'cat': 'brecha'
    },
    {
        'n': 'vinculacion_fecha_inconsistente',
        't': 'Vinculacion con fecha anterior al tamizaje',
        'd': 'La fecha de vinculacion es anterior a la fecha de tamizaje. '
             'No es posible vincular antes de tamizar.',
        'c': ['vinculo_fecha', 'fecha_tamizaje'],
        'resaltar': 'vinculo_fecha',
        'p': 'alta',
        'cat': 'consistencia'
    },
    {
        'n': 'poblacion_incoherente',
        't': 'Poblacion clave incoherente con sexo/edad',
        'd': 'Registros donde la poblacion clave no corresponde con '
             'el sexo biologico o la edad del paciente. '
             'Ej: HSH con sexo Femenino, Gestante con sexo Masculino, '
             'TS/TRA en menor de 15 anios.',
        'c': ['sexo', 'tipo_poblacion', 'edad'],
        'resaltar': 'tipo_poblacion',
        'p': 'alta',
        'cat': 'consistencia'
    },
    {
        'n': 'prep_sin_tamizaje_vih_reciente',
        't': 'PrEP sin tamizaje VIH en los ultimos 3 meses',
        'd': 'Paciente en PrEP (con condicion_actual) cuya fecha del ultimo '
             'tamizaje VIH es mayor a 90 dias o esta vacia.',
        'c': ['condicion_actual', 'fecha_ult_tamizaje_vih'],
        'resaltar': 'condicion_actual',
        'p': 'alta',
        'cat': 'consistencia'
    },
]

# ──────────────────────────────────────────────
# FUNCIONES DE VERIFICACION
# ──────────────────────────────────────────────
CRITERIOS = {v['n']: f"Columna(s): {', '.join(v['c'])}" for v in VERIFICACIONES}

RUTA_PLANTILLA = 'Trama_Unificada_encabezados.xlsx'


def _filtro_vih_dvi(df):
    """Filtra solo registros VIH y DVI (Dual VIH).
    Devuelve los que pasan el filtro para que se muestren en el reporte
    con la columna tipo_tamizaje resaltada en amarillo."""
    if 'tipo_tamizaje' not in df.columns:
        return pd.DataFrame()
    d = df[df['tipo_tamizaje'].str.upper().isin(['VIH', 'DVI'])]
    return d if not d.empty else pd.DataFrame()


def _duplicados_tamizaje(df):
    """Detectar duplicados por UID"""
    if 'uid' not in df.columns:
        return pd.DataFrame()
    dup_mask = df.duplicated(subset=['uid'], keep=False)
    d = df[dup_mask].copy()
    return d if not d.empty else pd.DataFrame()


def _reactivos_sin_vinculacion(df):
    """Reactivos VIH/Dual sin vinculacion efectiva"""
    if 'resultado' not in df.columns or 'vinculo_estado' not in df.columns:
        return pd.DataFrame()
    d = df[
        (df['resultado'].str.upper() == 'REACTIVO') &
        (~df['vinculo_estado'].str.upper().str.contains('EFECTIVA|EFECTIVO', na=False))
    ]
    return d if not d.empty else pd.DataFrame()


def _vinculados_sin_padron(df):
    """Vinculacion efectiva pero sin dato en TAR ni PrEP"""
    req = ['vinculo_estado', 'condicion_vih', 'condicion_actual']
    if not all(c in df.columns for c in req):
        return pd.DataFrame()
    efec = df['vinculo_estado'].str.upper().str.contains('EFECTIVA|EFECTIVO', na=False)
    sin_tar = df['condicion_vih'].isna() | (df['condicion_vih'] == '')
    sin_prep = df['condicion_actual'].isna() | (df['condicion_actual'] == '')
    d = df[efec & sin_tar & sin_prep]
    return d if not d.empty else pd.DataFrame()


def _reactivos_vih_sin_tar(df):
    """Reactivo VIH sin inicio de TAR"""
    if 'resultado' not in df.columns or 'fecha_inicio_tar' not in df.columns:
        return pd.DataFrame()
    d = df[
        (df['resultado'].str.upper() == 'REACTIVO') &
        (df['fecha_inicio_tar'].isna() | (df['fecha_inicio_tar'] == ''))
    ]
    return d if not d.empty else pd.DataFrame()


def _en_tar_sin_tamizaje_its(df):
    """En TAR pero sin tamizaje ITS"""
    if 'uid' not in df.columns or 'fecha_tamizaje' not in df.columns:
        return pd.DataFrame()
    d = df[
        (df['condicion_vih'].notna()) &
        (df['condicion_vih'] != '') &
        (df['fecha_tamizaje'].isna())
    ]
    return d if not d.empty else pd.DataFrame()


def _vinculacion_fecha_inconsistente(df):
    """Vinculacion antes del tamizaje"""
    if 'vinculo_fecha' not in df.columns or 'fecha_tamizaje' not in df.columns:
        return pd.DataFrame()
    d = df[
        df['vinculo_fecha'].notna() &
        df['fecha_tamizaje'].notna()
    ].copy()
    d['_vf'] = pd.to_datetime(d['vinculo_fecha'], errors='coerce')
    d['_ft'] = pd.to_datetime(d['fecha_tamizaje'], errors='coerce')
    d = d[d['_vf'] < d['_ft']]
    d = d.drop(columns=['_vf', '_ft'], errors='ignore')
    return d if not d.empty else pd.DataFrame()


def _poblacion_incoherente(df):
    """Poblacion clave vs sexo/edad"""
    if 'sexo' not in df.columns or 'tipo_poblacion' not in df.columns:
        return pd.DataFrame()
    d = df.copy()
    if 'edad' in d.columns:
        d['edad'] = pd.to_numeric(d['edad'], errors='coerce')
        menor = (d['edad'] < 15) & (d['tipo_poblacion'].str.upper().str.contains('TS|TRA|TRANS', na=False))
    else:
        menor = pd.Series(False, index=d.index)
    hsh_mujer = (
        (d['tipo_poblacion'].str.upper().str.contains('HSH', na=False)) &
        (d['sexo'].str.upper().str.contains('FEMENINO|FEMENINA|MUJER', na=False))
    )
    gestante_hombre = (
        (d['tipo_poblacion'].str.upper().str.contains('GESTANTE|PG|EMBARAZADA', na=False)) &
        (d['sexo'].str.upper().str.contains('MASCULINO|HOMBRE|VARON', na=False))
    )
    mask = menor | hsh_mujer | gestante_hombre
    d = d[mask]
    return d if not d.empty else pd.DataFrame()


def _prep_sin_tamizaje_vih_reciente(df):
    """PrEP activo sin tamizaje VIH en ultimos 3 meses"""
    if 'condicion_actual' not in df.columns or 'fecha_ult_tamizaje_vih' not in df.columns:
        return pd.DataFrame()
    d = df[
        (df['condicion_actual'].notna()) &
        (df['condicion_actual'] != '')
    ].copy()
    d['_futv'] = pd.to_datetime(d['fecha_ult_tamizaje_vih'], errors='coerce')
    hoy = pd.Timestamp.now()
    d['_dias'] = (hoy - d['_futv']).dt.days if not d['_futv'].isna().all() else 999
    d = d[d['_futv'].isna() | (d['_dias'] > 90)]
    d = d.drop(columns=['_futv', '_dias'], errors='ignore')
    return d if not d.empty else pd.DataFrame()


# ──────────────────────────────────────────────
# FUNCION MAPA: nombre -> funcion
# ──────────────────────────────────────────────
FUNCIONES = {
    'filtro_vih_dvi': _filtro_vih_dvi,
    'duplicados_tamizaje': _duplicados_tamizaje,
    'reactivos_sin_vinculacion': _reactivos_sin_vinculacion,
    'vinculados_sin_padron': _vinculados_sin_padron,
    'reactivos_vih_sin_tar': _reactivos_vih_sin_tar,
    'en_tar_sin_tamizaje_its': _en_tar_sin_tamizaje_its,
    'vinculacion_fecha_inconsistente': _vinculacion_fecha_inconsistente,
    'poblacion_incoherente': _poblacion_incoherente,
    'prep_sin_tamizaje_vih_reciente': _prep_sin_tamizaje_vih_reciente,
}


# ──────────────────────────────────────────────
# ANALIZADOR PRINCIPAL
# ──────────────────────────────────────────────
class AnalizadorRevisionBases:
    """Analiza la Trama Unificada (ITS + TAR + PrEP) para Revision de Bases"""

    MAPA_COLUMNAS = {
        'codigo_paciente': 'uid',
        'paciente_sexo': 'sexo',
        'paciente_edad': 'edad',
        'paciente_tipo_poblacion': 'tipo_poblacion',
        'resultado_1': 'resultado',
        'tipo_de_vinculacion': 'vinculo_tipo',
        'estado_de_vinculacion': 'vinculo_estado',
        'fecha_de_vinculacion': 'vinculo_fecha',
        'fecha_de_registro': 'vinculo_fecha_reg',
        'eess_destino_de_vinculacion': 'vinculo_eess_dest',
        'condicion': 'condicion_vih',
        'fecha_de_inicio_de_tar': 'fecha_inicio_tar',
        'fecha_de_ultima_atencion': 'fecha_ult_atencion',
        'fecha_de_proxima_cita': 'fecha_prox_cita',
        'fecha_de_ultimo_abandono': 'fecha_ult_abandono',
        'fecha_de_ultima_recuperacion_de_abandono': 'fecha_recup_abandono',
        'fecha_de_derivacion': 'fecha_derivacion',
        'lugar_de_destino_de_derivacion': 'lugar_derivacion',
        'eess_de_inicio_prep': 'eess_inicio_prep',
        'modalidad_de_inicio_prep': 'modalidad_prep',
        'fecha_de_reingreso_a_la_prep': 'fecha_reingreso_prep',
        'fecha_de_ultima_atencion_prep': 'fecha_ult_atencion_prep',
        'fecha_de_resultado_ultimo_tamizaje_de_vih': 'fecha_ult_tamizaje_vih',
        'resultado_ultimo_tamizaje_de_vih': 'resultado_ult_tamizaje_vih',
        'condicion_actual': 'condicion_actual',
        'sector___its': 'sector_its',
        'diresa___its': 'diresa_its',
        'region___its': 'region_its',
        'provincia___its': 'provincia_its',
        'distrito___its': 'distrito_its',
        'eess___its': 'eess_its',
        'codigo_renipres___its': 'renipres_its',
        'tipo_oferta': 'tipo_oferta_its',
        'tipo_lugar_intervencion': 'tipo_lugar_its',
        'fecha_intervencion': 'fecha_intervencion',
        'fecha_de_tamizaje': 'fecha_tamizaje',
        'tipo_tamizaje': 'tipo_tamizaje',
        'fecha_de_registro___its': 'fecha_reg_its',
        'fecha_de_modificacion___its': 'fecha_mod_its',
        'origen_de_registro': 'origen_registro',
        'sector___vih': 'sector_vih',
        'diris_diresa_geresa___vih': 'diris_vih',
        'departamento_eess___vih': 'depto_vih',
        'provincia_eess___vih': 'provincia_vih',
        'distrito_eess___vih': 'distrito_vih',
        'establecimiento_de_salud___vih': 'eess_vih',
        'codigo_renipres___vih': 'renipres_vih',
        'sector___prep': 'sector_prep',
        'diris_diresa_geresa___prep': 'diris_prep',
        'departamento_eess___prep': 'depto_prep',
        'provincia_eess___prep': 'provincia_prep',
        'distrito_eess___prep': 'distrito_prep',
        'establecimiento_de_salud___prep': 'eess_prep',
        'codigo_renipres___prep': 'renipres_prep',
        'fecha_de_creacion___vih': 'fecha_creacion_vih',
        'fecha_de_modificacion___vih': 'fecha_mod_vih',
        'fecha_de_creacion___prep': 'fecha_creacion_prep',
        'fecha_de_modificacion___prep': 'fecha_mod_prep',
        'fecha_de_inicio_de_prep': 'fecha_inicio_prep',
    }

    ENCABEZADOS = {
        # ITS
        'sector_its': 'SECTOR - ITS',
        'diresa_its': 'DIRESA - ITS',
        'region_its': 'REGION - ITS',
        'provincia_its': 'PROVINCIA - ITS',
        'distrito_its': 'DISTRITO - ITS',
        'eess_its': 'EESS - ITS',
        'renipres_its': 'CODIGO RENIPRES - ITS',
        'uid': 'CODIGO UID',
        'edad': 'PACIENTE EDAD',
        'sexo': 'PACIENTE SEXO',
        'tipo_poblacion': 'PACIENTE TIPO POBLACION',
        'nacionalidad': 'NACIONALIDAD',
        'etnia': 'ETNIA',
        'brigada': 'BRIGADA',
        'fecha_intervencion': 'FECHA INTERVENCION',
        'tipo_lugar_its': 'TIPO LUGAR INTERVENCION',
        'tipo_oferta_its': 'TIPO OFERTA',
        'fecha_tamizaje': 'FECHA DE TAMIZAJE',
        'tipo_tamizaje': 'TIPO TAMIZAJE',
        'resultado': 'RESULTADO 1',
        'fecha_reg_its': 'FECHA DE REGISTRO - ITS',
        'fecha_mod_its': 'FECHA DE MODIFICACION - ITS',
        'vinculo_tipo': 'TIPO DE VINCULACION',
        'vinculo_estado': 'ESTADO DE VINCULACION',
        'vinculo_eess_dest': 'EESS DESTINO DE VINCULACION',
        'vinculo_fecha': 'FECHA DE VINCULACION',
        'vinculo_fecha_reg': 'FECHA DE REGISTRO',
        'origen_registro': 'ORIGEN DE REGISTRO',
        # TAR
        'sector_vih': 'SECTOR - VIH',
        'diris_vih': 'DIRIS/DIRESA/GERESA - VIH',
        'depto_vih': 'DEPARTAMENTO EESS - VIH',
        'provincia_vih': 'PROVINCIA EESS - VIH',
        'distrito_vih': 'DISTRITO EESS - VIH',
        'eess_vih': 'ESTABLECIMIENTO DE SALUD - VIH',
        'renipres_vih': 'CODIGO RENIPRES - VIH',
        'condicion_vih': 'CONDICION',
        'fecha_inicio_tar': 'FECHA DE INICIO DE TAR',
        'fecha_ult_atencion': 'FECHA DE ULTIMA ATENCION',
        'fecha_prox_cita': 'FECHA DE PROXIMA CITA',
        'fecha_ult_abandono': 'FECHA DE ULTIMO ABANDONO',
        'fecha_recup_abandono': 'FECHA DE ULTIMA RECUPERACION DE ABANDONO',
        'fecha_derivacion': 'FECHA DE DERIVACION',
        'lugar_derivacion': 'LUGAR DE DESTINO DE DERIVACION',
        'fecha_creacion_vih': 'FECHA DE CREACION - VIH',
        'fecha_mod_vih': 'FECHA DE MODIFICACION - VIH',
        # PrEP
        'sector_prep': 'SECTOR - PREP',
        'diris_prep': 'DIRIS/DIRESA/GERESA - PREP',
        'depto_prep': 'DEPARTAMENTO EESS - PREP',
        'provincia_prep': 'PROVINCIA EESS - PREP',
        'distrito_prep': 'DISTRITO EESS - PREP',
        'eess_prep': 'ESTABLECIMIENTO DE SALUD - PREP',
        'renipres_prep': 'CODIGO RENIPRES - PREP',
        'fecha_inicio_prep': 'FECHA DE INICIO DE PREP',
        'eess_inicio_prep': 'EESS DE INICIO PREP',
        'modalidad_prep': 'MODALIDAD DE INICIO PREP',
        'fecha_reingreso_prep': 'FECHA DE REINGRESO A LA PREP',
        'condicion_actual': 'CONDICION ACTUAL',
        'fecha_ult_atencion_prep': 'FECHA DE ULTIMA ATENCION PREP',
        'fecha_ult_tamizaje_vih': 'FECHA DE RESULTADO ULTIMO TAMIZAJE DE VIH',
        'resultado_ult_tamizaje_vih': 'RESULTADO ULTIMO TAMIZAJE DE VIH',
        'fecha_creacion_prep': 'FECHA DE CREACION - PREP',
        'fecha_mod_prep': 'FECHA DE MODIFICACION - PREP',
    }

    def __init__(self):
        self.df = None
        self.total_registros = 0
        self.resumen = []
        self.resultados = {}
        self._v_omitidas = []
        self.tiempos = {}

    def cargar_datos(self, ruta_excel, hoja=None):
        """Carga el Excel de la Trama Unificada"""
        if hoja is None:
            try:
                xl = pd.ExcelFile(ruta_excel)
                hoja = xl.sheet_names[0]
            except Exception:
                pass
        df = pd.read_excel(ruta_excel, sheet_name=hoja)
        # Normalizar nombres de columnas
        df.columns = [c.strip().lower().replace(' ', '_').replace('-', '_') for c in df.columns]
        # Mapear a nombres canonicos
        rename_map = {}
        for col in df.columns:
            if col in self.MAPA_COLUMNAS:
                rename_map[col] = self.MAPA_COLUMNAS[col]
        if rename_map:
            df = df.rename(columns=rename_map)
            log.info(f"Columnas mapeadas: {len(rename_map)}")
        self.df = df
        self.total_registros = len(df)
        log.info(f"Cargados {self.total_registros} registros")
        return self

    def _campos_existen(self, cols):
        """Verifica que las columnas necesarias existan"""
        faltan = [c for c in cols if c not in self.df.columns]
        return faltan if faltan else None

    def _crear_hoja_con_formato(self, ws, registros, nombre_verif):
        """Crea hoja usando plantilla de encabezados, datos desde fila 4"""
        # Si no hay plantilla cargada, escribir datos simple
        if not getattr(self, '_wb_plantilla', None):
            rename_hdr = {k: v for k, v in self.ENCABEZADOS.items() if k in registros.columns}
            df = registros.copy()
            if rename_hdr:
                df = df.rename(columns=rename_hdr)
            for r_idx, (_, row) in enumerate(df.iterrows()):
                for c_idx, val in enumerate(row):
                    ws.cell(row=1 + r_idx, column=c_idx + 1, value=val)
            return ws

        ws_plant = self._wb_plantilla.worksheets[0]
        max_col = ws_plant.max_column

        # Construir set de celdas que pertenecen a un merge
        merged_cells_set = set()
        merge_top_left = {}  # (r,c) -> (tl_r, tl_c) para cada merge
        for mr in ws_plant.merged_cells.ranges:
            if mr.min_row <= 3:
                try:
                    ws.merge_cells(start_row=mr.min_row, start_column=mr.min_col,
                                   end_row=mr.max_row, end_column=mr.max_col)
                except Exception:
                    pass
                for r in range(mr.min_row, mr.max_row + 1):
                    for c in range(mr.min_col, mr.max_col + 1):
                        if r == mr.min_row and c == mr.min_col:
                            continue  # la top-left se escribe normal
                        merged_cells_set.add((r, c))
                    merge_top_left[(mr.min_row, mr.min_col)] = (mr.min_row, mr.min_col)

        # Copiar valores y estilos de filas 1-3
        for row in range(1, 4):
            for col in range(1, max_col + 1):
                if (row, col) in merged_cells_set:
                    continue
                src = ws_plant.cell(row=row, column=col)
                dst = ws.cell(row=row, column=col)
                dst.value = src.value
                if src.has_style:
                    try:
                        dst.font = src.font.copy()
                        dst.fill = src.fill.copy()
                        dst.alignment = src.alignment.copy()
                    except Exception:
                        pass

        # Colores fijos para fila 3 segun rangos
        # A-J (1-10) -> #4472C4, K-P (11-16) -> #FFC000
        # Q-Y (17-25) -> #4472C4, Z-AG (26-33) -> #FFC000
        # AH-BA (34-53) -> #ED7D31, BB-BU (54-73) -> #70AD47
        colores_f3 = {
            (1, 10): '4472C4',
            (11, 16): 'FFC000',
            (17, 25): '4472C4',
            (26, 33): 'FFC000',
            (34, 53): 'ED7D31',
            (54, 73): '70AD47',
        }
        fill_blanco = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
        font_f3 = Font(bold=True, color='FFFFFF', size=10)
        for inicio, fin in colores_f3:
            color = colores_f3[(inicio, fin)]
            fill_rango = PatternFill(start_color=color, end_color=color, fill_type='solid')
            for col in range(inicio, min(fin, max_col) + 1):
                if (3, col) not in merged_cells_set:
                    cell = ws.cell(row=3, column=col)
                    cell.fill = fill_rango
                    cell.font = font_f3
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # Copiar ancho de columnas
        for col_idx in range(1, max_col + 1):
            letter = get_column_letter(col_idx)
            if letter in ws_plant.column_dimensions:
                w = ws_plant.column_dimensions[letter].width
                if w:
                    ws.column_dimensions[letter].width = w

        # Renombrar columnas a encabezados legibles
        rename_hdr = {k: v for k, v in self.ENCABEZADOS.items() if k in registros.columns}
        df = registros.copy()
        if rename_hdr:
            df = df.rename(columns=rename_hdr)

        # Escribir datos desde fila 4
        for r_idx, (_, row) in enumerate(df.iterrows()):
            for c_idx, val in enumerate(row):
                cell = ws.cell(row=4 + r_idx, column=c_idx + 1)
                cell.value = val
                try:
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                except Exception:
                    pass

        return ws

    def analizar(self):
        """Ejecuta todas las verificaciones"""
        t_total = datetime.now()
        self.resumen = []
        self.resultados = {}

        log.info(f"Iniciando Revision de Bases sobre {self.total_registros} registros")

        # Paso 1: Ejecutar filtro VIH/DVI primero (siempre)
        log.info("  Aplicando filtro VIH/DVI antes del analisis...")
        filtro_fn = FUNCIONES.get('filtro_vih_dvi')
        df_filtrado = None
        if filtro_fn and 'tipo_tamizaje' in self.df.columns:
            df_filtrado = filtro_fn(self.df)
        if df_filtrado is not None and len(df_filtrado) > 0:
            log.info(f"  Filtro aplicado: {len(df_filtrado)} registros VIH/DVI de {self.total_registros} totales")
            self.df_original = self.df.copy()
            self.df = df_filtrado
            self.total_filtrados = len(df_filtrado)
        else:
            log.info("  No se pudo aplicar filtro VIH/DVI, se trabaja con todos los registros")
            self.df_original = None
            self.total_filtrados = self.total_registros

        for idx, item in enumerate(VERIFICACIONES, 1):
            t0 = pd.Timestamp.now()
            nombre = item['n']
            fn = FUNCIONES.get(nombre)

            faltan = self._campos_existen(item['c'])
            if faltan or fn is None:
                self._v_omitidas.append(nombre)
                detalle = f"Cols faltantes: {faltan}" if faltan else f"No implementada: {nombre}"
                self.resumen.append({
                    'ID': idx, 'Verificacion': item['t'],
                    'Categoria': item['cat'], 'Prioridad': item['p'],
                    'Registros': self.total_registros, 'Cantidad': 'N/A',
                    '%': 'N/A', 'Estado': 'OMITIDO',
                    'Criterio': CRITERIOS.get(nombre, ''),
                    'Descripcion': item['d'], 'Detalle': detalle,
                })
                log.warning(f"  OMITIDO: {nombre} -> {detalle}")
                continue

            try:
                # La verificacion #1 (filtro) usa df original, las demas usan df filtrado
                df_para_verif = self.df_original if nombre == 'filtro_vih_dvi' and hasattr(self, 'df_original') and self.df_original is not None else self.df
                registros = fn(df_para_verif)
                elapsed = (pd.Timestamp.now() - t0).total_seconds()
                self.tiempos[nombre] = elapsed

                total_base = len(df_para_verif)
                if registros is not None and len(registros) > 0:
                    cant = len(registros)
                    pct = (cant / total_base) * 100 if total_base > 0 else 0
                    self.resultados[nombre] = {
                        'id': idx, 'cantidad': cant, 'descripcion': item['d'],
                        'registros': registros, 'revisados': total_base,
                        'porcentaje': pct,
                    }
                    self.resumen.append({
                        'ID': idx, 'Verificacion': item['t'],
                        'Categoria': item['cat'], 'Prioridad': item['p'],
                        'Registros': total_base, 'Cantidad': cant,
                        '%': f"{pct:.2f}%", 'Estado': 'REVISAR',
                        'Criterio': CRITERIOS.get(nombre, ''),
                        'Descripcion': item['d'],
                    })
                    log.info(f"  -> {cant}/{total_base} registros ({pct:.2f}%) en {elapsed:.2f}s")
                else:
                    self.resumen.append({
                        'ID': idx, 'Verificacion': item['t'],
                        'Categoria': item['cat'], 'Prioridad': item['p'],
                        'Registros': total_base, 'Cantidad': 0,
                        '%': '0.00%', 'Estado': 'SIN PROBLEMAS',
                        'Criterio': CRITERIOS.get(nombre, ''),
                        'Descripcion': item['d'],
                    })
                    log.info(f"  -> 0 problemas en {elapsed:.2f}s")
            except Exception as e:
                elapsed = (pd.Timestamp.now() - t0).total_seconds()
                log.error(f"Error en {nombre}: {e}")
                self.resumen.append({
                    'ID': idx, 'Verificacion': item['t'],
                    'Categoria': item['cat'], 'Prioridad': item['p'],
                    'Registros': self.total_registros, 'Cantidad': 'ERROR',
                    '%': 'N/A', 'Estado': 'ERROR',
                    'Criterio': CRITERIOS.get(nombre, ''),
                    'Descripcion': item['d'], 'Detalle': str(e),
                })

        delta = (datetime.now() - t_total).total_seconds()
        log.info(f"Revision completa en {delta:.2f}s")
        return self

    def generar_reporte(self, nombre_base=None):
        """Genera Excel con resultados usando plantilla de encabezados"""
        if nombre_base is None:
            nombre_base = f"Revision_Bases_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        df_resumen = pd.DataFrame(self.resumen)

        # Cargar plantilla para copiar formato en hojas detalladas
        try:
            self._wb_plantilla = load_workbook(RUTA_PLANTILLA)
        except Exception as e:
            log.warning(f"No se pudo cargar plantilla '{RUTA_PLANTILLA}': {e}")
            self._wb_plantilla = None

        # Crear workbook base con solo resumenes
        with pd.ExcelWriter(nombre_base, engine='openpyxl') as writer:
            # Hoja Resumen General
            total_problemas = sum(
                r.get('Problemas', 0) for r in self.resumen
                if isinstance(r.get('Problemas', 0), (int, float))
            )
            criticos = sum(1 for r in self.resumen
                           if str(r.get('Prioridad', '')).lower() == 'critica'
                           and isinstance(r.get('Problemas', 0), (int, float))
                           and r.get('Problemas', 0) > 0)
            altos = sum(1 for r in self.resumen
                        if str(r.get('Prioridad', '')).lower() == 'alta'
                        and isinstance(r.get('Problemas', 0), (int, float))
                        and r.get('Problemas', 0) > 0)

            df_gen = pd.DataFrame({
                'Metrica': [
                    'Total Registros', 'Verificaciones Registradas',
                    'Verificaciones con Problemas', 'Verificaciones Omitidas',
                    'Total Problemas Encontrados', 'Criticos', 'Alta Prioridad',
                    'Fecha Analisis', 'Script'
                ],
                'Valor': [
                    self.total_registros, len(VERIFICACIONES),
                    sum(1 for r in self.resumen
                        if isinstance(r.get('Problemas', 0), (int, float))
                        and r.get('Problemas', 0) > 0),
                    len(self._v_omitidas),
                    total_problemas, criticos, altos,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'Revision Bases v1.0'
                ]
            })
            df_gen.to_excel(writer, sheet_name='Resumen General', index=False)

            # Hoja Resumen Verificaciones
            df_resumen.to_excel(writer, sheet_name='Resumen Verificaciones', index=False)

        # Cerrar para poder reabrir con openpyxl y agregar hojas con formato plantilla
        # (Usamos openpyxl directamente para las hojas detalladas)
        wb = load_workbook(nombre_base)

        # Hojas detalladas con formato de plantilla
        amarillo = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for nombre, res in self.resultados.items():
            registros = res.get('registros')
            if registros is None or not isinstance(registros, pd.DataFrame) or registros.empty:
                continue
            v = next((x for x in VERIFICACIONES if x['n'] == nombre), None)
            titulo = v['t'] if v else nombre
            safe_titulo = re.sub(r'[\\/*?:\[\]]', '_', titulo)[:27]
            sn = f"{res['id']:02d}_{safe_titulo}"

            # Crear hoja con formato plantilla + datos
            ws = wb.create_sheet(title=sn)
            self._crear_hoja_con_formato(ws, registros, nombre)

            # Resaltar columna clave en amarillo
            col_resaltar = v['resaltar'] if v and 'resaltar' in v else None
            if col_resaltar:
                col_excel = self.ENCABEZADOS.get(col_resaltar, col_resaltar)
                header = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
                if col_excel in header:
                    col_idx = header.index(col_excel) + 1
                    for row in range(4, ws.max_row + 1):
                        ws.cell(row=row, column=col_idx).fill = amarillo
                    log.info(f"Columna '{col_excel}' resaltada en hoja '{sn}'")

        wb.save(nombre_base)
        wb.close()

        log.info(f"Reporte final: {nombre_base}")
        return nombre_base
