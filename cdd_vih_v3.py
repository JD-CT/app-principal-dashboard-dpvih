# -*- coding: utf-8 -*-
# CDD VIH v3.0 — Calidad del Dato: TAR
import gc
import logging
import os
import re
import sys
import time
from datetime import datetime, date

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuration

class Config:
    
    COLUMNAS_FECHA= (
        'fecha_nacimiento', 'fecha_inicio_tar', 'fecha_fallecimiento',
        'fecha_abandono', 'fecha_derivacion', 'fecha_recuperacion_abandono',
        'fecha_esquema_actual', 'fecha_resultado_cv_basal',
        'prox_cita', 'fecha_resultado',
    )
    SIMILITUD_THRESHOLD= 90
    COLUMNAS_REQUERIDAS= (
        'nombres', 'apellido_paterno', 'tipo_documento', 'numero_documento'
    )
    SHEET_NAME= 'BD'
    TAR_MIN_YEAR= 2004
    EDAD_GESTACIONAL_MIN= 7
    EDAD_GESTACIONAL_MAX= 40
    PC_FEMENINO= ('PG', 'TS', 'PPL')
    PC_MASCULINO= ('PG', 'HSH', 'TRA', 'HTS', 'TTS', 'TS', 'PPL')
    PC_MENOR_15= ('PG', 'PPL')
    LONGITUDES_DOC= {
            'DNI': (8, 'DNI debe tener 8 digitos'),
            'carnet de extranjeria': (9, 'Carnet de extranjeria debe tener 9 digitos'),
            'di del extranjero': ((8, 12), 'DI del extranjero debe tener entre 8 y 12 digitos'),
            'pasaporte': ((8, 10), 'Pasaporte debe tener entre 8 y 10 caracteres'),
        }

# Estilos Excel

FILL_YELLOW = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
FILL_LIGHT_RED = PatternFill(start_color='FFD7D7', end_color='FFD7D7', fill_type='solid')
FILL_LIGHT_GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FONT_BOLD = Font(bold=True)
FONT_HEADER = Font(bold=True, color='FFFFFF')
FILL_HEADER = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')

# Logging estructurado

log = logging.getLogger('cdd_vih')

# Registry + Decorators

log = logging.getLogger('cdd_vih')

def _campos_existen(campos, df):
    faltantes = [c for c in campos if c not in df.columns]
    if faltantes:
        return f"Campos faltantes: {', '.join(faltantes)}"
    return None

    # Decorador que auto-registra cada verificación

def anotar_inconsistencia(df_resultado, columnas,
                          descripcion):
    if df_resultado.empty:
        return df_resultado
    df_resultado = df_resultado.copy()
    df_resultado.insert(0, 'Inconsistencia', descripcion)
    return df_resultado

# Lista plana de verificaciones (sin decoradores)

VERIFICACIONES = [
    {'n': 'nombres_similares', 'd': 'Registros con nombres similares (posibles duplicados)', 'c': ['nombres', 'apellido_paterno', 'apellido_materno'], 'p': 'alta', 'cat': 'duplicados'},
    {'n': 'duplicados_nombres', 'd': 'Nombres y apellidos duplicados exactos', 'c': ['nombres', 'apellido_paterno', 'apellido_materno'], 'p': 'alta', 'cat': 'duplicados'},
    {'n': 'documentos_duplicados', 'd': 'Mismo número doc. con distinto tipo', 'c': ['tipo_documento', 'numero_documento'], 'p': 'alta', 'cat': 'duplicados'},
    {'n': 'documentos_incorrectos', 'd': 'Documentos con longitud incorrecta según tipo', 'c': ['tipo_documento', 'numero_documento'], 'p': 'alta', 'cat': 'consistencia'},
    {'n': 'extranjeros_peru', 'd': 'Extranjeros con país de origen Perú', 'c': ['tipo_documento', 'pais_origen'], 'p': 'media', 'cat': 'consistencia'},
    {'n': 'tipo_doc_vacio', 'd': 'Tipo de documento no especificado', 'c': ['tipo_documento'], 'p': 'alta', 'cat': 'completitud'},
    {'n': 'pacientes_nuevos_incorrectos', 'd': 'Un solo registro no marcado como Nuevo', 'c': ['condicion', 'paciente_id'], 'p': 'alta', 'cat': 'reporte'},
    {'n': 'continuadores_incorrectos', 'd': 'Continuadores con < 2 atenciones', 'c': ['condicion', 'atenciones'], 'p': 'alta', 'cat': 'reporte'},
    {'n': 'derivados_sin_fecha', 'd': 'Derivado sin fecha / fecha sin derivado', 'c': ['condicion', 'fecha_derivacion'], 'p': 'media', 'cat': 'consistencia'},
    {'n': 'abandono_incorrecto', 'd': 'Fecha abandono sin condición', 'c': ['condicion', 'fecha_abandono', 'fecha_recuperacion_abandono'], 'p': 'alta', 'cat': 'consistencia'},
    {'n': 'abandono_por_prox_cita', 'd': 'Posible abandono: próxima cita vencida o ausente', 'c': ['condicion', 'prox_cita', 'fecha_abandono', 'fecha_recuperacion_abandono'], 'p': 'alta', 'cat': 'abandono'},
    {'n': 'fallecidos_sin_fecha', 'd': 'Fallecidos sin fecha de fallecimiento', 'c': ['condicion', 'fecha_fallecimiento'], 'p': 'alta', 'cat': 'completitud'},
    {'n': 'fechas_nacimiento_sospechosas', 'd': 'Fechas nac. sospechosas (peruanos)', 'c': ['fecha_nacimiento', 'pais_origen'], 'p': 'media', 'cat': 'consistencia'},
    {'n': 'extranjeros_fechas_sospechosas', 'd': 'Fechas nac. sospechosas (extranjeros)', 'c': ['fecha_nacimiento', 'pais_origen'], 'p': 'media', 'cat': 'consistencia'},
    {'n': 'nacimiento_posterior_tar', 'd': 'Fecha nac. posterior a inicio TAR', 'c': ['fecha_nacimiento', 'fecha_inicio_tar'], 'p': 'critica', 'cat': 'consistencia'},
    {'n': 'sexo_no_definido', 'd': 'Sexo biológico no definido o N/D', 'c': ['sexo_biologico'], 'p': 'alta', 'cat': 'completitud'},
    {'n': 'poblacion_incoherente', 'd': 'Incoherencia sexo ↔ población clave', 'c': ['sexo_biologico', 'poblacion_clave'], 'p': 'alta', 'cat': 'consistencia'},
    {'n': 'pais_vacio', 'd': 'País de origen no especificado', 'c': ['pais_origen'], 'p': 'baja', 'cat': 'completitud'},
    {'n': 'edad_gestacional_incorrecta', 'd': 'Edad gestacional fuera de rango', 'c': ['edad_gestacional'], 'p': 'media', 'cat': 'consistencia'},
    {'n': 'inicio_tar_antes_2004', 'd': 'Inicio TAR anterior a 2004', 'c': ['fecha_inicio_tar'], 'p': 'media', 'cat': 'consistencia'},
    {'n': 'primer_esquema_vacio', 'd': 'Primer esquema de tratamiento vacío', 'c': ['primer_esquema'], 'p': 'alta', 'cat': 'completitud'},
    {'n': 'fecha_esquema_actual_vacia', 'd': 'Fecha último esquema vacía', 'c': ['fecha_esquema_actual'], 'p': 'alta', 'cat': 'completitud'},
    {'n': 'tar_posterior_esquema', 'd': 'Inicio TAR posterior a fecha esquema', 'c': ['fecha_inicio_tar', 'fecha_esquema_actual'], 'p': 'alta', 'cat': 'consistencia'},
    {'n': 'esquemas_incorrectos', 'd': 'Esquema incompatible con fecha de inicio', 'c': ['esquema_actual', 'fecha_inicio_tar'], 'p': 'alta', 'cat': 'consistencia'},
    {'n': 'fallecimiento_mayor_100', 'd': 'Edad al fallecer > 100 años', 'c': ['fecha_fallecimiento', 'fecha_nacimiento'], 'p': 'media', 'cat': 'consistencia'},
    {'n': 'abandono_mayor_100', 'd': 'Edad al abandonar > 100 años', 'c': ['fecha_abandono', 'fecha_nacimiento'], 'p': 'media', 'cat': 'consistencia'},
    {'n': 'recuperacion_posterior_fallecimiento', 'd': 'Recuperación tras fallecimiento', 'c': ['fecha_recuperacion_abandono', 'fecha_fallecimiento'], 'p': 'critica', 'cat': 'consistencia'},
    {'n': 'carga_viral_anterior_tar', 'd': 'CV basal anterior a inicio TAR', 'c': ['fecha_resultado_cv_basal', 'fecha_inicio_tar'], 'p': 'alta', 'cat': 'consistencia'},
    {'n': 'ultimo_esquema_vacio', 'd': 'Último cambio de esquema vacío', 'c': ['ultimo_cambio_esquema'], 'p': 'alta', 'cat': 'completitud'},
    {'n': 'fecha_posterior_hoy', 'd': 'Fechas futuras (posteriores a hoy)', 'c': ['fecha_nacimiento', 'fecha_inicio_tar', 'fecha_fallecimiento', 'fecha_abandono', 'fecha_derivacion', 'fecha_recuperacion_abandono', 'fecha_esquema_actual', 'fecha_resultado_cv_basal', 'prox_cita', 'fecha_resultado'], 'p': 'alta', 'cat': 'consistencia'},
    {'n': 'esquema_actual_vs_cv', 'd': 'CV suprimida con esquema potencialmente incorrecto', 'c': ['esquema_actual', 'carga_viral', 'fecha_resultado'], 'p': 'media', 'cat': 'consistencia'},
    {'n': 'gestante_inconsistente', 'd': 'Gestante (sexo femenino, edad gestacional presente) sin PG', 'c': ['sexo_biologico', 'edad_gestacional', 'poblacion_clave'], 'p': 'alta', 'cat': 'consistencia'},
]

# Funciones de verificacion (planas, sin decoradores)

def _nombres_similares(df) :
    from rapidfuzz import fuzz
    tmp = df.copy()
    tmp['_nc'] = (tmp['nombres'].astype(str).str.strip().str.lower() + ' '
                  + tmp['apellido_paterno'].astype(str).str.strip().str.lower() + ' '
                  + tmp['apellido_materno'].astype(str).str.strip().str.lower())
    tmp['_nc'] = (tmp['_nc'].str.normalize('NFKD').str.encode('ascii', errors='ignore')
                   .str.decode('utf-8').str.replace(r'[^a-z0-9 ]', '', regex=True)
                   .str.replace(r'\s+', ' ', regex=True).str.strip())
    nombres_list = tmp['_nc'].tolist()
    idx_list = tmp.index.tolist()
    n = len(nombres_list)
    if n > 500:
        log.info('nombres_similares: %d registros, %d pares a revisar...', n, n*(n-1)//2)
    similares = []
    for i in range(n):
        for j in range(i + 1, n):
            score = fuzz.token_sort_ratio(nombres_list[i], nombres_list[j])
            if score >= 90:
                similares.append({
                    'idx1': idx_list[i],
                    'idx2': idx_list[j],
                    'nombre1': nombres_list[i],
                    'nombre2': nombres_list[j],
                    'similitud': score
                })
    tmp.drop('_nc', axis=1, inplace=True)
    if not similares:
        return pd.DataFrame()
    regs = []
    for r in similares:
        r1, r2 = df.loc[r['idx1']].copy(), df.loc[r['idx2']].copy()
        r1['Par Similar'] = f'Similar a registro {r["idx2"]}'
        r2['Par Similar'] = f'Similar a registro {r["idx1"]}'
        r1['Porcentaje Similitud'] = r['similitud']
        r2['Porcentaje Similitud'] = r['similitud']
        regs.extend([r1, r2])
    res = pd.concat(regs, axis=1).T if regs else pd.DataFrame()
    return anotar_inconsistencia(res, ['nombres','apellido_paterno','apellido_materno'],
                                 'Nombres y apellidos similares a otro registro (posible duplicado)')

def _duplicados_nombres(df):
    d = df[df.duplicated(subset=['nombres','apellido_paterno','apellido_materno'], keep=False)]
    return anotar_inconsistencia(d, ['nombres','apellido_paterno','apellido_materno'],
                                 "Mismos nombres y apellidos que otro registro")

def _documentos_duplicados(df):
    g = df.groupby('numero_documento')['tipo_documento'].nunique()
    d = df[df['numero_documento'].isin(g[g > 1].index)]
    return anotar_inconsistencia(d, ['tipo_documento','numero_documento'],
                                 "Mismo número con diferente tipo")

def _documentos_incorrectos(df):
    df = df.copy()
    if df['numero_documento'].dtype != 'object':
        df['numero_documento'] = df['numero_documento'].astype(str)
    cfg_doc = Config().LONGITUDES_DOC
    conds = []
    for tipo, (longitud, _) in cfg_doc.items():
        if isinstance(longitud, tuple):
            cond = ((df['tipo_documento'] == tipo)
                    & ((df['numero_documento'].str.len() < longitud[0])
                       | (df['numero_documento'].str.len() > longitud[1])))
        else:
            cond = ((df['tipo_documento'] == tipo)
                    & (df['numero_documento'].str.len() != longitud))
        conds.append(cond)
    mask = pd.concat(conds, axis=1).any(axis=1) if conds else pd.Series(False, index=df.index)
    incorrectos = df[mask].copy()
    incorrectos['_problema'] = ''
    for tipo, (_, mensaje) in cfg_doc.items():
        incorrectos.loc[incorrectos['tipo_documento'] == tipo, '_problema'] = mensaje
    return anotar_inconsistencia(incorrectos, ['tipo_documento','numero_documento'],
                                 "Longitud incorrecta del documento")

def _extranjeros_peru(df):
    docs_ext = ['carnet de extranjeria', 'di del extranjero', 'pasaporte']
    d = df[df['tipo_documento'].str.lower().isin(docs_ext)
           & (df['pais_origen'].str.lower() == 'peru')]
    return anotar_inconsistencia(d, ['tipo_documento','pais_origen'],
                                 "Extranjero con país Perú")

def _tipo_doc_vacio(df):
    d = df[df['tipo_documento'].isna() | (df['tipo_documento'].astype(str).str.strip() == '')]
    return anotar_inconsistencia(d, ['tipo_documento'], "Tipo de documento no especificado")

def _pacientes_nuevos_incorrectos(df):
    c = df.groupby('paciente_id').size().reset_index(name='ct')
    u = c[c['ct'] == 1]['paciente_id'].tolist()
    d = df[df['paciente_id'].isin(u) & (df['condicion'].str.upper() != 'NUEVO')]
    return anotar_inconsistencia(d, ['condicion','paciente_id'],
                                 "Un solo registro no marcado como NUEVO")

def _continuadores_incorrectos(df):
    d = df[(df['condicion'].str.upper() == 'CONTINUADOR') & (df['atenciones'] < 2)]
    return anotar_inconsistencia(d, ['condicion','atenciones'],
                                 "CONTINUADOR con menos de 2 atenciones")

def _derivados_sin_fecha(df):
    d1 = df[(df['condicion'].str.upper() == 'DERIVADO') & (df['fecha_derivacion'].isna())].copy()
    d1['_tipo'] = 'Derivado sin fecha'
    d2 = df[(df['fecha_derivacion'].notna()) & (df['condicion'].str.upper() != 'DERIVADO')].copy()
    d2['_tipo'] = 'Fecha sin derivado'
    return anotar_inconsistencia(pd.concat([d1,d2]), ['condicion','fecha_derivacion'],
                                 "Inconsistencia DERIVADO vs fecha derivación")

def _abandono_incorrecto(df):
    d = df[df['fecha_abandono'].notna() & df['fecha_recuperacion_abandono'].isna()
           & (df['condicion'].str.upper() != 'ABANDONO')]
    return anotar_inconsistencia(d, ['condicion','fecha_abandono','fecha_recuperacion_abandono'],
                                 "Fecha abandono sin condición ABANDONO")

# Abandono por próxima cita

def _abandono_por_prox_cita(df):
    
    hoy = date.today()
    df = df.copy()
    if 'prox_cita' in df.columns:
        df['prox_cita'] = pd.to_datetime(df['prox_cita'], errors='coerce')
    base = df[~df['condicion'].str.upper().isin(['FALLECIDO','DERIVADO','ABANDONO'])].copy()
    sin_recup = base['fecha_recuperacion_abandono'].isna()
    sin_cita = base['prox_cita'].isna()
    con_cita = base['prox_cita'].notna()
    # Diferencia en dias vectorizada
    dias_vencidos = (pd.Timestamp.now().normalize() - base['prox_cita']).dt.days
    vencida_larga = con_cita & (dias_vencidos > 30)
    sospechosos = base[(sin_cita | vencida_larga) & sin_recup].copy()
    # Clasificacion vectorizada
    sospechosos['_alerta_abandono'] = 'CRÍTICO: Sin próxima cita'
    mask_con_cita = sospechosos['prox_cita'].notna()
    sospechosos.loc[mask_con_cita, '_alerta_abandono'] = 'ALTO: ' + (pd.Timestamp.now().normalize() - sospechosos.loc[mask_con_cita, 'prox_cita']).dt.days.astype(str) + ' días sin cita'
    return anotar_inconsistencia(sospechosos, ['condicion','prox_cita','fecha_abandono'],
                                 "Posible abandono por cita vencida/ausente")

def _fallecidos_sin_fecha(df):
    d = df[(df['condicion'].str.upper() == 'FALLECIDO') & (df['fecha_fallecimiento'].isna())]
    return anotar_inconsistencia(d, ['condicion','fecha_fallecimiento'],
                                 "FALLECIDO sin fecha de fallecimiento")

def _fechas_nacimiento_sospechosas(df):
    d = df[(df['pais_origen'].str.lower() == 'peru')
           & (df['fecha_nacimiento'].isna()
              | (df['fecha_nacimiento'].dt.year == 1900)
              | (df['fecha_nacimiento'].dt.year == 1990))]
    return anotar_inconsistencia(d, ['fecha_nacimiento','pais_origen'],
                                 "Fecha nac. sospechosa (1900/1990/vacía) peruano")

def _extranjeros_fechas_sospechosas(df):
    d = df[(df['pais_origen'].str.lower() != 'peru')
           & (df['fecha_nacimiento'].isna()
              | (df['fecha_nacimiento'].dt.year == 1900)
              | (df['fecha_nacimiento'].dt.year == 1990))]
    return anotar_inconsistencia(d, ['fecha_nacimiento','pais_origen'],
                                 "Fecha nac. sospechosa (1900/1990/vacía) extranjero")

def _nacimiento_posterior_tar(df):
    d = df[df['fecha_nacimiento'].notna() & df['fecha_inicio_tar'].notna()
           & (df['fecha_nacimiento'] > df['fecha_inicio_tar'])]
    return anotar_inconsistencia(d, ['fecha_nacimiento','fecha_inicio_tar'],
                                 "Fecha nac. posterior a inicio TAR")

def _sexo_no_definido(df):
    d = df[df['sexo_biologico'].isna() | (df['sexo_biologico'].astype(str).str.upper() == 'N/D')]
    return anotar_inconsistencia(d, ['sexo_biologico'], "Sexo biológico no definido")

def _poblacion_incoherente(df):
    cfg = Config()
    fem = df[(df['sexo_biologico'].str.lower() == 'femenino')
             & df['poblacion_clave'].notna() & ~df['poblacion_clave'].isin(cfg.PC_FEMENINO)]
    masc = df[(df['sexo_biologico'].str.lower() == 'masculino')
              & df['poblacion_clave'].notna() & ~df['poblacion_clave'].isin(cfg.PC_MASCULINO)]
    r = pd.concat([fem, masc])
    if 'edad_actual' in df.columns:
        men = df[(df['sexo_biologico'].str.lower() == 'masculino') & (df['edad_actual'] < 15)
                 & df['poblacion_clave'].notna() & ~df['poblacion_clave'].isin(cfg.PC_MENOR_15)]
        r = pd.concat([r, men])
    return anotar_inconsistencia(r, ['sexo_biologico','poblacion_clave'],
                                 "Incoherencia sexo ↔ población clave")

def _pais_vacio(df):
    return anotar_inconsistencia(df[df['pais_origen'].isna()], ['pais_origen'],
                                 "País de origen vacío")

def _edad_gestacional_incorrecta(df):
    cfg = Config()
    d = df[df['edad_gestacional'].notna()
           & ((df['edad_gestacional'] < cfg.EDAD_GESTACIONAL_MIN)
              | (df['edad_gestacional'] > cfg.EDAD_GESTACIONAL_MAX))]
    return anotar_inconsistencia(d, ['edad_gestacional'],
                                 f"Edad gestacional fuera de {cfg.EDAD_GESTACIONAL_MIN}-{cfg.EDAD_GESTACIONAL_MAX} sem")

def _inicio_tar_antes_2004(df):
    cfg = Config()
    d = df[df['fecha_inicio_tar'].notna() & (df['fecha_inicio_tar'].dt.year < cfg.TAR_MIN_YEAR)]
    return anotar_inconsistencia(d, ['fecha_inicio_tar'],
                                 f"Inicio TAR anterior a {cfg.TAR_MIN_YEAR}")

def _primer_esquema_vacio(df):
    return anotar_inconsistencia(df[df['primer_esquema'].isna()], ['primer_esquema'],
                                 "Primer esquema no especificado")

def _fecha_esquema_actual_vacia(df):
    return anotar_inconsistencia(df[df['fecha_esquema_actual'].isna()], ['fecha_esquema_actual'],
                                 "Fecha último esquema no especificada")

def _tar_posterior_esquema(df):
    d = df[df['fecha_inicio_tar'].notna() & df['fecha_esquema_actual'].notna()
           & (df['fecha_inicio_tar'] > df['fecha_esquema_actual'])]
    return anotar_inconsistencia(d, ['fecha_inicio_tar','fecha_esquema_actual'],
                                 "Inicio TAR posterior a fecha último esquema")

def _esquemas_incorrectos(df):
    condiciones = [
        (df['esquema_actual'].str.contains('TDF/3TC/DTG', na=False)
         & df['fecha_inicio_tar'].notna() & (df['fecha_inicio_tar'].dt.year < 2019),
         "TDF/3TC/DTG con inicio TAR < 2019"),
        (df['esquema_actual'].str.contains(r'TDF/3TC/EFV\\$400\\$', na=False, regex=True)
         & df['fecha_inicio_tar'].notna() & (df['fecha_inicio_tar'].dt.year < 2019),
         "TDF/3TC/EFV 400 con inicio TAR < 2019"),
        (df['esquema_actual'].str.contains('DDI', na=False)
         & df['fecha_inicio_tar'].notna() & (df['fecha_inicio_tar'].dt.year > 2015),
         "DDI con inicio TAR > 2015 (descontinuado)"),
        (df['esquema_actual'].str.contains('NVP', na=False)
         & df['fecha_inicio_tar'].notna() & (df['fecha_inicio_tar'].dt.year > 2020),
         "NVP con inicio TAR > 2020"),
        (df['esquema_actual'].str.contains('D4T', na=False)
         & df['fecha_inicio_tar'].notna() & (df['fecha_inicio_tar'].dt.year > 2015),
         "D4T con inicio TAR > 2015"),
        (df['esquema_actual'].str.contains('IDV', na=False)
         & df['fecha_inicio_tar'].notna() & (df['fecha_inicio_tar'].dt.year > 2008),
         "IDV con inicio TAR > 2008"),
        (df['primer_esquema'].str.contains('SQV', na=False)
         & df['fecha_inicio_tar'].notna() & (df['fecha_inicio_tar'].dt.year > 2014),
         "SQV como primer esquema con inicio > 2014"),
        (df['primer_esquema'].str.contains('NFV', na=False)
         & df['fecha_inicio_tar'].notna() & (df['fecha_inicio_tar'].dt.year > 2015),
         "NFV como primer esquema con inicio > 2015"),
    ]
    partes = []
    for cond, desc in condiciones:
        sub = df[cond].copy()
        if len(sub):
            sub['_detalle_esquema'] = desc
            partes.append(sub)
    r = pd.concat(partes) if partes else pd.DataFrame()
    return anotar_inconsistencia(r, ['esquema_actual','fecha_inicio_tar'],
                                 "Esquema incompatible con fecha inicio TAR")

def _fallecimiento_mayor_100(df):
    df = df.copy()
    df['_edad_fall'] = (df['fecha_fallecimiento'] - df['fecha_nacimiento']).dt.days / 365.25
    d = df[df['fecha_fallecimiento'].notna() & df['fecha_nacimiento'].notna()
           & (df['_edad_fall'] > 100)]
    df.drop('_edad_fall', axis=1, inplace=True)
    return anotar_inconsistencia(d, ['fecha_fallecimiento','fecha_nacimiento'],
                                 "Edad al fallecer > 100 años")

def _abandono_mayor_100(df):
    df = df.copy()
    df['_edad_ab'] = (df['fecha_abandono'] - df['fecha_nacimiento']).dt.days / 365.25
    d = df[df['fecha_abandono'].notna() & df['fecha_nacimiento'].notna()
           & (df['_edad_ab'] > 100)]
    df.drop('_edad_ab', axis=1, inplace=True)
    return anotar_inconsistencia(d, ['fecha_abandono','fecha_nacimiento'],
                                 "Edad al abandonar > 100 años")

def _recuperacion_posterior_fallecimiento(df):
    d = df[df['fecha_recuperacion_abandono'].notna() & df['fecha_fallecimiento'].notna()
           & (df['fecha_recuperacion_abandono'] > df['fecha_fallecimiento'])]
    return anotar_inconsistencia(d, ['fecha_recuperacion_abandono','fecha_fallecimiento'],
                                 "Recuperación posterior a fallecimiento")

def _carga_viral_anterior_tar(df):
    d = df[df['fecha_resultado_cv_basal'].notna() & df['fecha_inicio_tar'].notna()
           & (df['fecha_resultado_cv_basal'] < df['fecha_inicio_tar'])]
    return anotar_inconsistencia(d, ['fecha_resultado_cv_basal','fecha_inicio_tar'],
                                 "CV anterior a inicio TAR")

def _ultimo_esquema_vacio(df):
    return anotar_inconsistencia(df[df['ultimo_cambio_esquema'].isna()],
                                 ['ultimo_cambio_esquema'],
                                 "Último cambio de esquema no especificado")

# Fechas posteriores a hoy

def _fecha_posterior_hoy(df):
    hoy = pd.Timestamp.now().normalize()
    cfg = Config()
    partes = []
    for col in cfg.COLUMNAS_FECHA:
        if col not in df.columns:
            continue
        col_dt = pd.to_datetime(df[col], errors='coerce')
        futuras = df[col_dt.notna() & (col_dt > hoy)].copy()
        if len(futuras):
            futuras['_col_futura'] = col
            partes.append(futuras)
    r = pd.concat(partes) if partes else pd.DataFrame()
    return anotar_inconsistencia(r, list(cfg.COLUMNAS_FECHA[:4]),
                                 "Fechas posteriores a hoy (posible error de captura)")

# Esquema actual vs carga viral (CV suprimida pero esquema antiguo)

def _esquema_actual_vs_cv(df):
    
    if 'carga_viral' not in df.columns:
        return pd.DataFrame()
    df = df.copy()
    if 'fecha_resultado' in df.columns:
        df['fecha_resultado'] = pd.to_datetime(df['fecha_resultado'], errors='coerce')
    drogas_descontinuadas = 'DDI|D4T|IDV'
    d = df[
        df['esquema_actual'].str.contains(drogas_descontinuadas, na=False, regex=True)
        & (df['carga_viral'].notna())
    ]
    # Si tenemos resultado de CV numérica, filtrar suprimidas < 1000
    if pd.api.types.is_numeric_dtype(d['carga_viral']):
        d = d[d['carga_viral'] < 1000]
    return anotar_inconsistencia(d, ['esquema_actual','carga_viral','fecha_resultado'],
                                 "CV suprimida con esquema descontinuado")

# Gestante sin población clave PG

def _gestante_inconsistente(df):
    d = df[
        (df['sexo_biologico'].fillna('').str.lower() == 'femenino')
        & df['edad_gestacional'].notna()
        & (df['edad_gestacional'] > 0)
        & (df['poblacion_clave'].fillna('') != 'PG')
    ]
    return anotar_inconsistencia(d, ['sexo_biologico','edad_gestacional','poblacion_clave'],
                                 "Gestante con edad gestacional pero sin PG")

# ANALIZADOR


# Mapa nombre_funcion -> funcion

_FN_MAP = {
    'nombres_similares': _nombres_similares,
    'duplicados_nombres': _duplicados_nombres,
    'documentos_duplicados': _documentos_duplicados,
    'documentos_incorrectos': _documentos_incorrectos,
    'extranjeros_peru': _extranjeros_peru,
    'tipo_doc_vacio': _tipo_doc_vacio,
    'pacientes_nuevos_incorrectos': _pacientes_nuevos_incorrectos,
    'continuadores_incorrectos': _continuadores_incorrectos,
    'derivados_sin_fecha': _derivados_sin_fecha,
    'abandono_incorrecto': _abandono_incorrecto,
    'abandono_por_prox_cita': _abandono_por_prox_cita,
    'fallecidos_sin_fecha': _fallecidos_sin_fecha,
    'fechas_nacimiento_sospechosas': _fechas_nacimiento_sospechosas,
    'extranjeros_fechas_sospechosas': _extranjeros_fechas_sospechosas,
    'nacimiento_posterior_tar': _nacimiento_posterior_tar,
    'sexo_no_definido': _sexo_no_definido,
    'poblacion_incoherente': _poblacion_incoherente,
    'pais_vacio': _pais_vacio,
    'edad_gestacional_incorrecta': _edad_gestacional_incorrecta,
    'inicio_tar_antes_2004': _inicio_tar_antes_2004,
    'primer_esquema_vacio': _primer_esquema_vacio,
    'fecha_esquema_actual_vacia': _fecha_esquema_actual_vacia,
    'tar_posterior_esquema': _tar_posterior_esquema,
    'esquemas_incorrectos': _esquemas_incorrectos,
    'fallecimiento_mayor_100': _fallecimiento_mayor_100,
    'abandono_mayor_100': _abandono_mayor_100,
    'recuperacion_posterior_fallecimiento': _recuperacion_posterior_fallecimiento,
    'carga_viral_anterior_tar': _carga_viral_anterior_tar,
    'ultimo_esquema_vacio': _ultimo_esquema_vacio,
    'fecha_posterior_hoy': _fecha_posterior_hoy,
    'esquema_actual_vs_cv': _esquema_actual_vs_cv,
    'gestante_inconsistente': _gestante_inconsistente,
}

def _ejecutar_verificacion(nombre, descripcion, campos, df):
    error = _campos_existen(campos, df)
    if error:
        return None, error
    fn = _FN_MAP.get(nombre)
    if fn is None:
        return None, f'Verificacion no implementada: {nombre}'
    return fn(df), None

class AnalizadorCalidadDatos:

    def __init__(self):
        self.df = None
        self.resultados= {}
        self.resumen = []
        self.tiempos= {}
        self.total_registros = 0
        self._verificaciones_omitidas = []

    def cargar_datos(self, ruta):
        if not os.path.exists(ruta):
            raise FileNotFoundError(f"Archivo no existe: {ruta}")
        log.info(f"Cargando: {ruta}")
        self.df = pd.read_excel(ruta, sheet_name=Config.SHEET_NAME)
        self.total_registros = len(self.df)
        if self.df.empty:
            raise ValueError("Archivo vacío (0 registros)")
        # Renombrar columnas legacy
        rename_map = {
            'esquema_tratamiento_actual': 'ultimo_cambio_esquema',
            'fecha_esquema_tratamiento': 'fecha_esquema_actual',
        }
        for old, new in rename_map.items():
            if old in self.df.columns:
                self.df.rename(columns={old: new}, inplace=True)
                log.info(f"Columna renombrada: {old} → {new}")
        self._convertir_fechas()
        self._validar_estructura()
        if 'condicion' in self.df.columns:
            self.df['condicion'] = self.df['condicion'].astype(str).str.strip().str.upper()
        log.info(f"Cargados {self.total_registros} registros, {len(self.df.columns)} columnas")

    def _convertir_fechas(self) :
        cfg = Config()
        for col in cfg.COLUMNAS_FECHA:
            if col in self.df.columns:
                self.df[col] = pd.to_datetime(self.df[col], errors='coerce')
            else:
                log.debug(f"Columna fecha no encontrada: {col}")

    def _validar_estructura(self) :
        cfg = Config()
        faltantes = [c for c in cfg.COLUMNAS_REQUERIDAS if c not in self.df.columns]
        if faltantes:
            raise ValueError(f"Columnas requeridas faltantes: {', '.join(faltantes)}")

    def analizar(self) :
        if self.df is None:
            raise ValueError("No hay datos cargados")
        log.info(f"Iniciando análisis: {self.total_registros} registros, "
                 f"{len(VERIFICACIONES)} verificaciones registradas")

        self.resultados.clear()
        self.resumen.clear()
        self.tiempos.clear()
        self._verificaciones_omitidas.clear()

        for idx, item in enumerate(VERIFICACIONES, 1):
            nombre = item['n']
            t0 = time.perf_counter()
            fn = _FN_MAP.get(nombre)
            log.info(f"[{idx}/{len(VERIFICACIONES)}] {item['d']}")

            # Verificar campos
            error = _campos_existen(item['c'], self.df)
            if error or fn is None:
                self._verificaciones_omitidas.append(nombre)
                self.resumen.append({
                    'ID': idx, 'Verificación': nombre.replace('_', ' ').title(),
                    'Categoría': item['cat'], 'Prioridad': item['p'],
                    'Registros': 0, 'Problemas': 'N/A', '%': 'N/A',
                    'Estado': 'OMITIDO', 'Detalle': error or f'No implementada: {nombre}',
                })
                continue

            try:
                registros_problema = fn(self.df)
                elapsed = time.perf_counter() - t0
                self.tiempos[nombre] = elapsed

                if registros_problema is not None and len(registros_problema) > 0:
                    cant = len(registros_problema)
                    pct = (cant / self.total_registros) * 100
                    self.resultados[nombre] = {
                        'id': idx, 'cantidad': cant, 'descripcion': item['d'],
                        'registros': registros_problema, 'revisados': self.total_registros,
                        'porcentaje': pct,
                    }
                    self.resumen.append({
                        'ID': idx, 'Verificación': nombre.replace('_', ' ').title(),
                        'Categoría': item['cat'], 'Prioridad': item['p'],
                        'Registros': self.total_registros, 'Problemas': cant,
                        '%': f"{pct:.2f}%", 'Estado': 'OK',
                        'Detalle': item['d'],
                    })
                    log.info(f"  → {cant} problemas ({pct:.2f}%) en {elapsed:.2f}s")
                else:
                    self.resumen.append({
                        'ID': idx, 'Verificación': nombre.replace('_', ' ').title(),
                        'Categoría': item['cat'], 'Prioridad': item['p'],
                        'Registros': self.total_registros, 'Problemas': 0,
                        '%': '0.00%', 'Estado': 'SIN PROBLEMAS',
                        'Detalle': item['d'],
                    })
                    log.info(f"  → 0 problemas en {elapsed:.2f}s")
            except Exception as e:
                elapsed = time.perf_counter() - t0
                log.error(f"Error en {nombre}: {str(e)}")
                self.resumen.append({
                    'ID': idx, 'Verificación': nombre.replace('_', ' ').title(),
                    'Categoría': item['cat'], 'Prioridad': item['p'],
                    'Registros': self.total_registros, 'Problemas': 'ERROR',
                    '%': 'N/A', 'Estado': 'ERROR', 'Detalle': str(e),
                })

            gc.collect()

    def generar_reporte(self, nombre_archivo=None):
        if not self.resumen:
            raise ValueError("No hay resultados")
        if nombre_archivo is None:
            nombre_archivo = f"CDD_VIH_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        wb = Workbook()
        wb.remove(wb.active)

        # Hoja 1: Resumen General
        ws = wb.create_sheet("Resumen General", 0)
        info = [
            ('Total Registros', self.total_registros),
            ('Verificaciones Registradas', len(VERIFICACIONES)),
            ('Verificaciones Ejecutadas', len([r for r in self.resumen if r['Estado'] == 'OK' or r['Estado'] == 'SIN PROBLEMAS'])),
            ('Verificaciones con Problemas', sum(1 for r in self.resumen if isinstance(r.get('Problemas'), int) and r['Problemas'] > 0)),
            ('Verificaciones Omitidas', len(self._verificaciones_omitidas)),
            ('Fecha Análisis', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Script', 'CDD VIH v3.0'),
        ]
        for r_idx, (k, v) in enumerate(info, 1):
            ws.cell(row=r_idx, column=1, value=k).font = FONT_BOLD
            ws.cell(row=r_idx, column=2, value=v)

        # Hoja 2: Resumen Verificaciones
        ws2 = wb.create_sheet("Resumen Verificaciones")
        df_res = pd.DataFrame(self.resumen)
        for r_idx, row in enumerate(dataframe_to_rows(df_res, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = FONT_HEADER
                    cell.fill = FILL_HEADER

        # Hoja 3: Tiempos
        ws3 = wb.create_sheet("Tiempos Ejecución")
        ws3.cell(row=1, column=1, value='Verificación').font = FONT_HEADER
        ws3.cell(row=1, column=1).fill = FILL_HEADER
        ws3.cell(row=1, column=2, value='Tiempo (s)').font = FONT_HEADER
        ws3.cell(row=1, column=2).fill = FILL_HEADER
        for r_idx, (k, v) in enumerate(sorted(self.tiempos.items(), key=lambda x: -x[1]), 2):
            ws3.cell(row=r_idx, column=1, value=k.replace('_', ' ').title())
            ws3.cell(row=r_idx, column=2, value=round(v, 2))

        # Hojas detalladas por verificación con problemas
        for meta in VERIFICACIONES:
            nom = meta['n']
            if nom not in self.resultados:
                continue
            res = self.resultados[nom]
            registros = res['registros']
            if registros.empty:
                continue

            sheet_name = f"{res['id']:02d}_{nom[:22]}"
            ws_det = wb.create_sheet(sheet_name)

            # Cabecera informativa
            ws_det.cell(row=1, column=1, value='VERIFICACIÓN').font = FONT_BOLD
            ws_det.cell(row=1, column=2, value=meta['d'])
            ws_det.cell(row=2, column=1, value='CAMPOS').font = FONT_BOLD
            ws_det.cell(row=2, column=2, value=', '.join(meta['c']))
            ws_det.cell(row=3, column=1, value='TOTAL REGISTROS').font = FONT_BOLD
            ws_det.cell(row=3, column=2, value=res['revisados'])
            ws_det.cell(row=4, column=1, value='PROBLEMAS').font = FONT_BOLD
            ws_det.cell(row=4, column=2, value=res['cantidad'])
            ws_det.cell(row=5, column=1, value='%').font = FONT_BOLD
            ws_det.cell(row=5, column=2, value=f"{res['porcentaje']:.2f}%")
            ws_det.cell(row=6, column=1, value='PRIORIDAD').font = FONT_BOLD
            ws_det.cell(row=6, column=2, value=meta['p'].upper())
            ws_det.cell(row=7, column=1, value='CATEGORÍA').font = FONT_BOLD
            ws_det.cell(row=7, column=2, value=meta['cat'].upper())

            # Datos
            start_row = 9
            campo_inc = 'Inconsistencia'
            cols_to_write = [campo_inc] if campo_inc in registros.columns else []
            cols_to_write += [c for c in registros.columns if c != campo_inc]

            for r_idx, row in enumerate(dataframe_to_rows(registros[cols_to_write], index=False, header=True), start_row):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_det.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == start_row:
                        cell.font = FONT_BOLD
                        if cols_to_write[c_idx-1] in meta['c']:
                            cell.fill = FILL_YELLOW
                    elif cols_to_write[c_idx-1] in meta['c']:
                        cell.fill = FILL_YELLOW

            # Autoajuste
            for col in ws_det.columns:
                max_len = max(len(str(c.value or '')) for c in col[:50])
                ws_det.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        wb.save(nombre_archivo)
        log.info(f"Reporte: {nombre_archivo}")
        return nombre_archivo

# CLI

def _listar_exceles() :
    return [f for f in os.listdir() if f.endswith(('.xlsx', '.xls'))]

def main():
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        handlers=[logging.FileHandler('cdd_vih_v3.log'), logging.StreamHandler()],
    )

    analizador = AnalizadorCalidadDatos()

    # Intentar argumento de línea
    if len(sys.argv) > 1:
        ruta = sys.argv[1]
    else:
        archivos = _listar_exceles()
        if archivos:
            print(f"Archivos: {len(archivos)}:")
            for i, f in enumerate(archivos, 1):
                print(f"  {i}. {f}")
        ruta = input("\nRuta del archivo Excel (Enter para salir): ").strip()
        if not ruta:
            print("Cancelado.")
            return

    ruta = ruta.strip()
    if not os.path.exists(ruta):
        print(f"Error: archivo no existe: {ruta}")
        sys.exit(1)

    analizador.cargar_datos(ruta)
    analizador.analizar()
    reporte = analizador.generar_reporte()

    print(f"\n{'='*60}")
    print(f"{'='*60}")
    print("\nResumen:")

    # Tabla bonita del resumen
    df_show = pd.DataFrame(analizador.resumen)
    print(df_show.to_string(index=False))

    # Recomendaciones
    criticos = [r for r in analizador.resumen
                if isinstance(r.get('Problemas'), int) and r['Problemas'] > 0
                and r.get('Prioridad') == 'critica']
    if criticos:
        for c in criticos:
            print(f"  - {c['Verificación']}: {c['Problemas']} registro(s) ({c['%']})")

if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrumpido por el usuario.")
        sys.exit(0)
    except Exception as e:
        log.error(f"Error fatal: {str(e)}")
        print(f"\nError: {e}")
        sys.exit(1)
