#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Escenarios de Vinculacion 2026 - Trama Unificada
# 3 validaciones con 3 sub-pasos cada una
# Basado en: proceso manual de 6 pasos → 3 escenarios automatizados

import pandas as pd
import numpy as np
from datetime import datetime
import logging, re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Logging ───
log = logging.getLogger('escenarios')

# ─── CONSTANTES ───
ANIO = 2026

# Mapeo de columnas de la Trama Unificada v1.6 (73 col)
TRAMA_COLUMNAS = {
    'uid': 'CÓDIGO UID',
    'documento_its': 'NÚMERO DE DOCUMENTO USUARIO ITS / BRIGADISTA',
    'documento_vih': 'NÚMERO DE DOCUMENTO USUARIO VIH',
    'documento_prep': 'NÚMERO DE DOCUMENTO USUARIO PREP',
    'edad': 'PACIENTE EDAD',
    'sexo': 'PACIENTE SEXO',
    'eess_its': 'EESS - ITS',
    'tamizaje_fecha': 'FECHA DE TAMIZAJE',
    'tamizaje_tipo': 'TIPO TAMIZAJE',
    'tamizaje_resultado': 'RESULTADO 1',
    'vinculo_tipo': 'TIPO DE VINCULACIÓN',
    'vinculo_estado': 'ESTADO DE VINCULACION',
    'vinculo_fecha': 'FECHA DE VINCULACION',
    'vinculo_eess_dest': 'EESS DESTINO DE VINCULACION',
    'eess_vih': 'ESTABLECIMIENTO DE SALUD - VIH',
    'fecha_inicio_tar': 'FECHA DE INICIO DE TAR',
    'fecha_ultima_atencion_vih': 'FECHA DE ÚLTIMA ATENCIÓN',
    'eess_prep': 'ESTABLECIMIENTO DE SALUD - PREP',
    'fecha_inicio_prep': 'FECHA DE INICIO DE PREP',
    'fecha_ultima_atencion_prep': 'FECHA DE ÚLTIMA ATENCIÓN PREP',
}

def _get_fecha(df, col):
    """Convierte columna a datetime si no lo es"""
    if col not in df.columns:
        return df
    if not pd.api.types.is_datetime64_any_dtype(df[col]):
        df[col] = pd.to_datetime(df[col], errors='coerce')
    return df

def _en_anio(df, col, anio=ANIO):
    """Filtra registros donde columna fecha cae en el anio indicado"""
    if col not in df.columns:
        return pd.Series(False, index=df.index)
    df = _get_fecha(df, col)
    return df[col].dt.year == anio

def _tiene_valor(df, col):
    """Retorna True si la columna no es nula"""
    if col not in df.columns:
        return pd.Series(False, index=df.index)
    return df[col].notna() & (df[col] != '')

def _num_doc(df):
    """Obtiene el numero de documento priorizando ITS > VIH > PREP"""
    doc = df.get('documento_its', pd.Series('', index=df.index)).fillna('')
    m1 = doc.str.len() >= 7
    doc_ok = doc.where(m1, df.get('documento_vih', pd.Series('', index=df.index)).fillna(''))
    m2 = doc_ok.str.len() >= 7
    doc_ok = doc_ok.where(m2, df.get('documento_prep', pd.Series('', index=df.index)).fillna(''))
    return doc_ok

# ─── VALIDACIÓN 01: Tamizajes vinculados efectivamente ───
def _v01_identificacion(df):
    """Paso 1: Tamizajes VIH/DUAL VIH con vinculacion efectiva"""
    es_vih = df.get('tamizaje_tipo', pd.Series('', index=df.index)).str.upper().str.contains('VIH', na=False)
    tiene_vinculo = _tiene_valor(df, 'vinculo_fecha') | _tiene_valor(df, 'vinculo_tipo')
    en_anio = _en_anio(df, 'tamizaje_fecha')
    cond = es_vih & tiene_vinculo & en_anio
    return df[cond].copy()

def _v01_validar_tar(df):
    """Paso 2: De los vinculados, validar que esten en TAR 2026"""
    en_tar = _en_anio(df, 'fecha_inicio_tar')
    return df[en_tar].copy()

def _v01_validar_prep(df):
    """Paso 3: De los vinculados, validar que esten en PREP 2026"""
    en_prep = _en_anio(df, 'fecha_inicio_prep')
    return df[en_prep].copy()

# ─── VALIDACIÓN 02: Tamizajes sin vinculacion efectiva ───
def _v02_identificacion(df):
    """Paso 1: Tamizajes VIH/DUAL VIH sin vinculacion efectiva"""
    es_vih = df.get('tamizaje_tipo', pd.Series('', index=df.index)).str.upper().str.contains('VIH', na=False)
    sin_vinculo = ~(_tiene_valor(df, 'vinculo_fecha') | _tiene_valor(df, 'vinculo_tipo'))
    en_anio = _en_anio(df, 'tamizaje_fecha')
    cond = es_vih & sin_vinculo & en_anio
    return df[cond].copy()

def _v02_validar_tar(df):
    """Paso 2: De los no vinculados, ver si aparecen en TAR 2026"""
    en_tar = _en_anio(df, 'fecha_inicio_tar')
    return df[en_tar].copy()

def _v02_validar_prep(df):
    """Paso 3: De los no vinculados, ver si aparecen en PREP 2026"""
    en_prep = _en_anio(df, 'fecha_inicio_prep')
    return df[en_prep].copy()

# ─── VALIDACIÓN 03: TAR/PrEP sin tamizaje registrado ───
def _v03_identificacion(df):
    """Paso 1: Pacientes en TAR/PrEP 2026 sin tamizaje VIH asociado"""
    en_tar_o_prep = _en_anio(df, 'fecha_inicio_tar') | _en_anio(df, 'fecha_inicio_prep')
    tiene_tamizaje = _tiene_valor(df, 'tamizaje_fecha')
    return df[en_tar_o_prep & ~tiene_tamizaje].copy()

def _v03_validar_tar(df):
    """Paso 2: De esos, confirmar que estan en TAR 2026"""
    en_tar = _en_anio(df, 'fecha_inicio_tar')
    return df[en_tar].copy()

def _v03_validar_prep(df):
    """Paso 3: De esos, confirmar que estan en PREP 2026"""
    en_prep = _en_anio(df, 'fecha_inicio_prep')
    return df[en_prep].copy()


# ─── REGISTRO DE VALIDACIONES ───
VALIDACIONES = [
    {
        'id': 'V01',
        'nombre': 'Tamizajes vinculados efectivamente',
        'pasos': [
            {'id': 'V01.1', 'nombre': 'Identificacion ITS', 'fn': _v01_identificacion},
            {'id': 'V01.2', 'nombre': 'Validacion en padron TAR 2026', 'fn': _v01_validar_tar},
            {'id': 'V01.3', 'nombre': 'Validacion en padron PrEP 2026', 'fn': _v01_validar_prep},
        ]
    },
    {
        'id': 'V02',
        'nombre': 'Tamizajes sin vinculacion efectiva',
        'pasos': [
            {'id': 'V02.1', 'nombre': 'Identificacion ITS', 'fn': _v02_identificacion},
            {'id': 'V02.2', 'nombre': 'Validacion en padron TAR 2026', 'fn': _v02_validar_tar},
            {'id': 'V02.3', 'nombre': 'Validacion en padron PrEP 2026', 'fn': _v02_validar_prep},
        ]
    },
    {
        'id': 'V03',
        'nombre': 'TAR y PrEP sin registro de tamizajes',
        'pasos': [
            {'id': 'V03.1', 'nombre': 'Identificacion ITS', 'fn': _v03_identificacion},
            {'id': 'V03.2', 'nombre': 'Validacion en padron TAR 2026', 'fn': _v03_validar_tar},
            {'id': 'V03.3', 'nombre': 'Validacion en padron PrEP 2026', 'fn': _v03_validar_prep},
        ]
    },
]

# Columnas a exportar en detalle
COLUMNAS_DETALLE = [
    'uid', 'edad', 'sexo',
    'eess_its', 'tamizaje_fecha', 'tamizaje_tipo', 'tamizaje_resultado',
    'vinculo_tipo', 'vinculo_estado', 'vinculo_fecha',
    'eess_vih', 'fecha_inicio_tar', 'fecha_ultima_atencion_vih',
    'eess_prep', 'fecha_inicio_prep', 'fecha_ultima_atencion_prep',
]

ENCABEZADOS_EXCEL = {
    'uid': 'CODIGO UID',
    'edad': 'EDAD',
    'sexo': 'SEXO',
    'eess_its': 'EESS ITS',
    'tamizaje_fecha': 'FECHA TAMIZAJE',
    'tamizaje_tipo': 'TIPO TAMIZAJE',
    'tamizaje_resultado': 'RESULTADO 1',
    'vinculo_tipo': 'TIPO VINCULACION',
    'vinculo_estado': 'ESTADO VINCULACION',
    'vinculo_fecha': 'FECHA VINCULACION',
    'eess_vih': 'EESS VIH',
    'fecha_inicio_tar': 'FECHA INICIO TAR',
    'fecha_ultima_atencion_vih': 'ULTIMA ATENCION VIH',
    'eess_prep': 'EESS PREP',
    'fecha_inicio_prep': 'FECHA INICIO PREP',
    'fecha_ultima_atencion_prep': 'ULTIMA ATENCION PREP',
}


class AnalizadorEscenarios:
    """Ejecuta las 3 validaciones de escenarios 2026"""

    def __init__(self):
        self.df = None
        self.resultados = {}
        self.resumen = []

    def cargar_datos(self, ruta_excel, hoja=None):
        """Carga la Trama Unificada desde Excel"""
        log.info(f"Cargando datos desde: {ruta_excel}")
        self.df = pd.read_excel(ruta_excel, sheet_name=hoja if hoja else 0)

        # Mapear columnas de la trama a nombres cortos
        invertido = {v: k for k, v in TRAMA_COLUMNAS.items()}
        cols_disponibles = set(self.df.columns)
        self.cols_mapeadas = {}
        self.cols_faltantes = []

        for col_trama, col_corta in invertido.items():
            # Buscar con strip() por espacios al final
            encontrada = None
            for col_real in cols_disponibles:
                if col_real.strip() == col_trama.strip():
                    encontrada = col_real
                    break
            if encontrada:
                self.cols_mapeadas[col_corta] = encontrada
            else:
                self.cols_faltantes.append(col_trama)

        if self.cols_faltantes:
            log.warning(f"Columnas faltantes ({len(self.cols_faltantes)}): {self.cols_faltantes[:5]}...")

        # Renombrar columnas existentes (usando el nombre original de la trama)
        rename_map = {v: k for k, v in self.cols_mapeadas.items()}
        self.df = self.df.rename(columns=rename_map)

        # Convertir fechas
        for col in ['tamizaje_fecha', 'vinculo_fecha', 'fecha_inicio_tar',
                     'fecha_ultima_atencion_vih', 'fecha_inicio_prep',
                     'fecha_ultima_atencion_prep']:
            self.df = _get_fecha(self.df, col)

        # Generar num_documento (con proteccion)
        try:
            self.df['num_documento'] = _num_doc(self.df)
        except Exception as e:
            log.warning(f"No se pudo generar num_documento: {e}")
            self.df['num_documento'] = ''

        log.info(f"Cargados {len(self.df)} registros con {len(self.cols_mapeadas)}/{len(TRAMA_COLUMNAS)} columnas")
        return self

    def analizar(self):
        """Ejecuta todas las validaciones"""
        if self.df is None:
            raise ValueError("Primero debe cargar datos con cargar_datos()")

        log.info("=== INICIANDO ANALISIS DE ESCENARIOS 2026 ===")
        for val in VALIDACIONES:
            log.info(f"\n--- {val['id']}: {val['nombre']} ---")
            paso_anterior_df = self.df.copy()

            for paso in val['pasos']:
                log.info(f"  Paso {paso['id']}: {paso['nombre']}...")
                try:
                    resultado = paso['fn'](paso_anterior_df)
                    if resultado is not None and len(resultado) > 0:
                        cant = len(resultado)
                        log.info(f"    -> {cant} registros")
                        self.resultados[paso['id']] = {
                            'validacion': val['id'],
                            'validacion_nombre': val['nombre'],
                            'paso': paso['nombre'],
                            'registros': resultado,
                            'cantidad': cant,
                        }
                    else:
                        log.info(f"    -> 0 registros")
                        self.resultados[paso['id']] = {
                            'validacion': val['id'],
                            'validacion_nombre': val['nombre'],
                            'paso': paso['nombre'],
                            'registros': pd.DataFrame(),
                            'cantidad': 0,
                        }
                except Exception as e:
                    log.error(f"  Error en {paso['id']}: {e}")
                    self.resultados[paso['id']] = {
                        'validacion': val['id'],
                        'validacion_nombre': val['nombre'],
                        'paso': paso['nombre'],
                        'registros': pd.DataFrame(),
                        'cantidad': 0,
                        'error': str(e),
                    }

        # Generar resumen
        self._generar_resumen()
        log.info("=== ANALISIS COMPLETADO ===")
        return self

    def _generar_resumen(self):
        """Construye el DataFrame de resumen"""
        filas = []
        for val in VALIDACIONES:
            for paso in val['pasos']:
                res = self.resultados.get(paso['id'], {})
                cant = res.get('cantidad', 0)
                total_paso = len(self.df) if paso['id'].endswith('.1') else (
                    self.resultados.get(val['pasos'][0]['id'], {}).get('cantidad', 0)
                )
                cumple = cant
                no_cumple = max(0, total_paso - cant)
                pct = (cumple / total_paso * 100) if total_paso > 0 else 0
                filas.append({
                    'Validacion': paso['id'],
                    'Validacion_Nombre': val['nombre'],
                    'Paso': paso['nombre'],
                    'Total_Pacientes': total_paso,
                    'Cumplen': cumple,
                    'No_Cumplen': no_cumple,
                    'Porcentaje': f"{pct:.1f}%",
                })
        self.resumen = pd.DataFrame(filas)

    def generar_excel(self, ruta_salida=None):
        """Genera Excel con resumen + detalle"""
        if ruta_salida is None:
            ruta_salida = f"Escenarios_Vinculacion_{ANIO}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        log.info(f"Generando Excel: {ruta_salida}")

        # --- Hoja Resumen ---
        with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
            self.resumen.to_excel(writer, sheet_name='Resumen', index=False)

            # --- Hoja Detalle ---
            filas_detalle = []
            for val in VALIDACIONES:
                for paso in val['pasos']:
                    res = self.resultados.get(paso['id'], {})
                    registros = res.get('registros', pd.DataFrame())
                    if registros is None or registros.empty:
                        continue

                    tmp = registros.copy()
                    # Solo columnas que existen
                    cols_exportar = [c for c in COLUMNAS_DETALLE if c in tmp.columns]
                    tmp = tmp[cols_exportar].copy()
                    tmp.insert(0, 'Paso', paso['id'])
                    tmp.insert(0, 'Validacion', val['id'])

                    # Renombrar para Excel
                    rename_excel = {k: v for k, v in ENCABEZADOS_EXCEL.items() if k in tmp.columns}
                    if rename_excel:
                        tmp = tmp.rename(columns=rename_excel)

                    # Fechas a solo fecha
                    for col in tmp.select_dtypes(include=['datetime64']).columns:
                        tmp[col] = tmp[col].dt.strftime('%Y-%m-%d')

                    filas_detalle.append(tmp)

            if filas_detalle:
                df_detalle = pd.concat(filas_detalle, ignore_index=True)
                df_detalle.to_excel(writer, sheet_name='Detalle', index=False)
            else:
                pd.DataFrame({
                    'Observacion': ['Sin datos para el periodo 2026']
                }).to_excel(writer, sheet_name='Detalle', index=False)

        # Formato con openpyxl
        wb = load_workbook(ruta_salida)

        # Formato Resumen
        if 'Resumen' in wb.sheetnames:
            ws = wb['Resumen']
            self._formatear_hoja(ws)

        # Formato Detalle
        if 'Detalle' in wb.sheetnames:
            ws = wb['Detalle']
            self._formatear_hoja(ws)

            # Resaltar filas donde Estado = REVISAR (si existe columna Estado)
            header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            if 'Estado' in header:
                col_estado = header.index('Estado') + 1
                rojo_fill = PatternFill(start_color='FFF5F5', end_color='FFF5F5', fill_type='solid')
                for row in range(2, ws.max_row + 1):
                    if str(ws.cell(row=row, column=col_estado).value).upper() == 'REVISAR':
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = rojo_fill

        wb.save(ruta_salida)
        wb.close()
        log.info(f"Excel generado: {ruta_salida}")
        return ruta_salida

    def _formatear_hoja(self, ws):
        """Formato basico de hoja"""
        # Encabezados en negrita
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
        header_font_white = Font(bold=True, size=10, color='FFFFFF')

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Autoancho
        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row in range(1, min(ws.max_row + 1, 50)):
                val = ws.cell(row=row, column=col).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 40)

        # Alternar colores de fila
        light_fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')
        for row in range(2, ws.max_row + 1):
            if row % 2 == 0:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = light_fill


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(name)s] %(message)s')
    import sys
    if len(sys.argv) < 2:
        print("Uso: python3 escenarios_vinculacion.py <trama_unificada.xlsx>")
        sys.exit(1)
    analizador = AnalizadorEscenarios()
    analizador.cargar_datos(sys.argv[1]).analizar()
    ruta = analizador.generar_excel()
    print(f"Excel generado: {ruta}")
